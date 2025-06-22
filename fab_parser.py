import time
import os
import argparse
import openpyxl
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from urllib.parse import urljoin


def get_fab_assets(url, output_filename="fab_assets.xlsx", chromedriver_path=None, append_mode=False):
    """
    Extracts asset names, links, prices/status, ratings, and rating counts from a Fab.com seller page.
    Uses Selenium for dynamic content loading (scrolling).
    Results are saved to an Excel file.

    Args:
        url (str): URL of the seller's page on Fab.com.
        output_filename (str): Name of the output Excel file (should end with .xlsx).
        chromedriver_path (str, optional): Full path to chromedriver.exe.
        append_mode (bool): If True, data will be appended to an existing file.
                            If False, the file will be overwritten.
    """

    driver = None
    try:
        # ChromeDriver setup
        if chromedriver_path:
            service = webdriver.chrome.service.Service(executable_path=chromedriver_path)
            driver = webdriver.Chrome(service=service)
        else:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            local_chromedriver_path = os.path.join(script_dir, 'chromedriver.exe')
            if os.path.exists(local_chromedriver_path):
                service = webdriver.chrome.service.Service(executable_path=local_chromedriver_path)
                driver = webdriver.Chrome(service=service)
            else:
                driver = webdriver.Chrome()

        print("ChromeDriver successfully initialized.")

        print(f"Opening page: {url}")
        driver.get(url)

        print("Attempting to close cookie pop-up (if present)...")
        try:
            cookie_accept_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
            )
            cookie_accept_button.click()
            print("Cookie pop-up closed.")
            time.sleep(2)
        except Exception:
            print("Cookie pop-up not found or not closed (possibly not present). Continuing.")

        # 3. Wait for the first set of assets to load (looking for div.fabkit-Stack-root)
        print("Waiting for the first set of assets to load (looking for div.fabkit-Stack-root)...")
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'fabkit-Stack-root'))
        )
        print("First set of assets loaded.")

        # 4. Scroll to load all assets
        print("Starting scroll to load all assets. This may take some time...")
        last_height = driver.execute_script("return document.body.scrollHeight")
        scroll_attempts = 0
        max_scroll_attempts = 10
        while scroll_attempts < max_scroll_attempts:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(3)

            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                print("Scrolling complete: all assets loaded.")
                break
            last_height = new_height
            scroll_attempts += 1
            print(
                f"Scrolled {scroll_attempts}/{max_scroll_attempts}. New page height: {new_height}px. Continuing scroll...")
        else:
            print(f"Maximum scroll attempts ({max_scroll_attempts}) reached. Not all assets might be loaded.")

        # 5. Get page HTML content after all loads
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')

        # 6. Parse HTML
        # We will always find all fabkit-Stack-root, but then filter them more strictly
        all_potential_asset_containers = soup.find_all('div', class_='fabkit-Stack-root')

        # Filter for actual product cards based on essential components
        actual_asset_containers = []
        print(
            f"Found {len(all_potential_asset_containers)} potential asset containers. Filtering for actual product cards...")
        for container in all_potential_asset_containers:
            product_link_tag = container.find('a')
            product_image_tag = container.find('img')
            product_title_div = container.find('div', class_='fabkit-Typography-ellipsisWrapper')

            # If these core elements exist, it's likely a real product card
            if product_link_tag and product_image_tag and product_title_div:
                actual_asset_containers.append(container)
            # else:
            #     # For debugging purposes:
            #     # print("Skipping potential container due to missing core product elements:")
            #     # print(container.prettify())

        if not actual_asset_containers:
            print(
                "Error: After filtering, no actual product asset containers were found. This might indicate a change in page structure or no products on the page.")
            driver.save_screenshot("debug_screenshot_no_actual_asset_containers.png")
            print("Screenshot 'debug_screenshot_no_actual_asset_containers.png' taken for debugging.")
            return

        current_url = driver.current_url
        unique_assets_data = set()  # Store tuples of (title, link, price_status, rating, rating_count)

        print(f"Found {len(actual_asset_containers)} actual product asset containers. Extracting unique data...")
        for i, asset_div in enumerate(actual_asset_containers):
            # Re-find these for data extraction within the confirmed product asset_div
            asset_link_tag = asset_div.find('a')

            title = None
            link = None
            price_status = None
            rating = None
            rating_count = None

            if asset_link_tag and asset_link_tag.get('href'):
                title_div = asset_link_tag.find('div', class_='fabkit-Typography-ellipsisWrapper')
                if title_div:
                    title = title_div.get_text(strip=True)

                link = asset_link_tag.get('href')

                if link and link.startswith('/'):
                    link = urljoin(url, link)

                    # --- Extract Sold Out Status --- (Priority 1)
                sold_out_container = asset_div.find('div',
                                                    class_='fabkit-Typography-root fabkit-Typography--align-start fabkit-Typography--intent-success fabkit-Text--sm fabkit-Text--regular fabkit-Stack-root fabkit-Stack--align_center fabkit-scale--gapX-spacing-1 fabkit-scale--gapY-spacing-1 dK8TLWWt')
                if sold_out_container:
                    # Confirm it has the checkmark icon as a child
                    if sold_out_container.find('i',
                                               class_='fabkit-Icon-root fabkit-Icon--intent-success fabkit-Icon--xs edsicon edsicon-check-circle-filled'):
                        price_status = sold_out_container.get_text(strip=True)  # Get text directly from this div

                # --- Extract Price --- (Priority 2, only if not Sold Out)
                if price_status is None:
                    # Parent Div 1 (grandparent of price text)
                    price_grandparent_div = asset_div.find('div',
                                                           class_='fabkit-Stack-root fabkit-Stack--align_center fabkit-scale--gapX-spacing-2 fabkit-scale--gapY-spacing-2 csZFzinF')
                    if price_grandparent_div:
                        # Parent Div 2 (parent of price text)
                        price_parent_div = price_grandparent_div.find('div',
                                                                      class_='fabkit-Stack-root fabkit-scale--gapX-spacing-1 fabkit-scale--gapY-spacing-1 J9vFXlBh')
                        if price_parent_div:
                            # Target Price Text Div
                            price_text_div = price_parent_div.find('div',
                                                                   class_='fabkit-Typography-root fabkit-Typography--align-start fabkit-Typography--intent-primary fabkit-Text--sm fabkit-Text--regular')
                            if price_text_div:
                                price_status = price_text_div.get_text(strip=True)

                # --- Extract Rating and Rating Count --- (Independent search within asset_div)
                # Parent Div (contains star icon and rating number)
                rating_main_div = asset_div.find('div',
                                                 class_='fabkit-Stack-root fabkit-Stack--align_center fabkit-scale--gapX-spacing-1 fabkit-scale--gapY-spacing-1')
                if rating_main_div:
                    # Check for the star icon to confirm this is the rating section
                    star_icon = rating_main_div.find('i',
                                                     class_='fabkit-Icon-root fabkit-Icon--intent-warning fabkit-Icon--xs edsicon edsicon-star-filled')
                    if star_icon:
                        # Rating number
                        rating_value_tag = rating_main_div.find('div',
                                                                class_='fabkit-Typography-root fabkit-Typography--align-start fabkit-Typography--intent-primary fabkit-Text--sm fabkit-Text--regular')
                        if rating_value_tag:
                            rating = rating_value_tag.get_text(strip=True)

                        # Rating count
                        rating_count_tag = rating_main_div.find('div',
                                                                class_='fabkit-Typography-root fabkit-Typography--align-start fabkit-Typography--intent-secondary fabkit-Text--sm fabkit-Text--regular')
                        if rating_count_tag:
                            rating_count = rating_count_tag.get_text(strip=True)

                # Add to set only if title and link are found (basic product info)
                if title and link:
                    # Use a tuple for the set to ensure uniqueness
                    unique_assets_data.add((title, link, price_status, rating, rating_count))

                    # 7. Write unique assets to Excel
        if unique_assets_data:
            workbook = None
            sheet = None
            start_row = 1

            if append_mode and os.path.exists(output_filename):
                print(f"Append mode: Loading existing file '{output_filename}'.")
                try:
                    workbook = openpyxl.load_workbook(output_filename)
                    sheet = workbook.active
                    start_row = sheet.max_row + 1
                    print(f"Data will be appended starting from row {start_row}.")
                except Exception as e:
                    print(f"Error loading existing Excel file: {e}. Creating a new file.")
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    sheet.title = "Fab Assets"
                    # Headers are added only if the file is new or if loading failed
                    sheet['A1'] = 'Product Name'
                    sheet['B1'] = 'Product Link'
                    sheet['C1'] = 'Price/Status'
                    sheet['D1'] = 'Rating'
                    sheet['E1'] = 'Rating Count'
                    start_row = 2
            else:
                print(f"Creating a new file '{output_filename}' (or overwriting existing).")
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Fab Assets"
                # Column Headers
                sheet['A1'] = 'Product Name'
                sheet['B1'] = 'Product Link'
                sheet['C1'] = 'Price/Status'
                sheet['D1'] = 'Rating'
                sheet['E1'] = 'Rating Count'
                start_row = 2

            row_num = start_row

            # Collect existing data from the file to avoid duplicates when appending
            existing_data = set()
            if append_mode and start_row > 2:  # If appending and headers are already present
                # Read existing data (Product Name, Product Link) for uniqueness check
                for row_data in sheet.iter_rows(min_row=2, values_only=True):
                    # Only consider the first two columns (Name, Link) for uniqueness
                    if len(row_data) >= 2 and row_data[0] and row_data[1]:
                        existing_data.add((str(row_data[0]), str(row_data[1])))

            newly_added_count = 0
            # Sort data before writing for readability
            sorted_assets = sorted(list(unique_assets_data), key=lambda x: x[0] if x[0] else '')
            for title, link, price_status, rating, rating_count in sorted_assets:
                # Add only if this record is not already in the file (checking by title and link)
                if (title, link) not in existing_data:
                    sheet.cell(row=row_num, column=1).value = title
                    sheet.cell(row=row_num, column=2).value = link
                    sheet.cell(row=row_num, column=3).value = price_status
                    sheet.cell(row=row_num, column=4).value = rating
                    sheet.cell(row=row_num, column=5).value = rating_count
                    row_num += 1
                    newly_added_count += 1

            # Auto-adjust column widths
            for col_idx in range(1, sheet.max_column + 1):
                max_length = 0
                column = get_column_letter(col_idx)
                for cell in sheet[column]:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                if adjusted_width > 50:  # Cap max width for very long text
                    adjusted_width = 50
                sheet.column_dimensions[column].width = adjusted_width

            # Save the Excel file
            workbook.save(output_filename)
            if append_mode and newly_added_count > 0:
                print(f"{newly_added_count} new unique assets appended to '{output_filename}'.")
            elif append_mode and newly_added_count == 0:
                print(f"No new unique assets found to append to '{output_filename}'.")
            else:  # Overwrite mode
                print(
                    f"Information on {len(unique_assets_data)} unique assets successfully saved to '{output_filename}' (file was overwritten).")
        else:
            print("Warning: No unique assets found to save.")


    except Exception as e:
        print(f"An error occurred: {e}")
        if driver:
            try:
                driver.save_screenshot("error_screenshot.png")
                print("Screenshot 'error_screenshot.png' taken for debugging.")
            except Exception as ss_e:
                print(f"Could not take screenshot: {ss_e}")
        import traceback
        traceback.print_exc()
    finally:
        if driver:
            print("Closing browser.")
            driver.quit()


if __name__ == "__main__":
    print("Starting script execution...")

    parser = argparse.ArgumentParser(
        description="Script to extract product names, links, prices, ratings, and rating counts from Fab.com and save them to Excel.")

    parser.add_argument("url", type=str,
                        help="URL of the seller's page on Fab.com (e.g., https://www.fab.com/sellers/NaughtyMonk)")

    parser.add_argument("-o", "--output", type=str, default="fab_assets.xlsx",
                        help="Name of the output Excel file (default: fab_assets.xlsx). Must end with .xlsx")

    parser.add_argument("-d", "--driver_path", type=str, default=None,
                        help="Full path to chromedriver.exe (if not in system PATH or next to the script).")

    parser.add_argument("-a", "--append", action="store_true",
                        help="Append data to an existing Excel file instead of overwriting it. If the file does not exist, it will be created.")

    args = parser.parse_args()

    # Ensure the output file has a .xlsx extension
    if not args.output.lower().endswith(".xlsx"):
        args.output += ".xlsx"
        print(f"Warning: Output filename changed to '{args.output}' for correct Excel format.")

    get_fab_assets(args.url, args.output, chromedriver_path=args.driver_path, append_mode=args.append)

    print("Script execution finished.")